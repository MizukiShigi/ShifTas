from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import login
from django.views.generic import TemplateView
from staff.models import Staff
from submitshifts.models import SubmitShift
from datetime import datetime,time
from common import common
import io
import openpyxl
from django.http import HttpResponse

class SubmitShifsView(TemplateView):
    template_name = 'shifts/shift_submit.html'
    def get(self, request):
        
        context = {}
        shifts = []
        dt_now = datetime.now()
        year = dt_now.year

        # 月の指定があればその月の表を出力
        if 'month' in request.GET:
            month_str = request.GET.get('month')
            month = int(month_str)
        else:
            month = dt_now.month + 1
        days = common.getCalendarDays(year, month)
        
        staffs = Staff.objects.all()
        staff_id = Staff.objects.get(staff=request.user)
        request_shift = SubmitShift.objects.filter(year=year, month=month, staff_id=staff_id)
        
        for staff in staffs:
            target_shifts = SubmitShift.objects.filter(year=year, month=month, staff_id=staff).order_by('day')
            shifts.append({'name':staff.staff_name, 'shifts':target_shifts})
        print(shifts)
        status = staff_id
        context = {
            'year'          :year, 
            'month'         :month, 
            'this_month'    :dt_now.month, 
            'days'          :days,
            'status'        :status,
            'staffs'        :staffs,
            'request_shift' :request_shift,
            'shifts_list'   :shifts,
            'range'         :range(1, len(days)+1),
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        staff_id = Staff.objects.get(staff=request.user)
        submit_year = request.POST.get('year')
        submit_month = request.POST.get('month')
        # 提出されたシフトで変更や新規登録があったものを更新、追加
        for req in request.POST:
            if 'request' in req:
                submit_day = req.strip('_request')
                request_shift = request.POST.get(req)
                print("request_shift")
                print(request_shift)
                # 更新処理
                try:
                    obj = SubmitShift.objects.get(staff_id=staff_id, year=submit_year, month=submit_month, day=submit_day)
                    if request_shift == 'x':
                        obj.absence_flg = True
                    elif request_shift != '':
                        request_fromtime = int(request_shift[:request_shift.find('-')])
                        request_totime = int(request_shift[request_shift.find('-')+1:])
                        obj.fromtime = request_fromtime
                        obj.totime = request_totime
                        obj.absence_flg = False
                    obj.save()
                # 追加処理
                except SubmitShift.DoesNotExist:
                    new_values = {'staff_id':staff_id, 'year':submit_year, 'month':submit_month, 'day':submit_day}
                    if request_shift == '' or request_shift == 'x':
                        new_values['absence_flg'] = True
                    elif request_shift != '':
                        request_fromtime = int(request_shift[:request_shift.find('-')])
                        request_totime = int(request_shift[request_shift.find('-')+1:])
                        new_values['fromtime'] = request_fromtime
                        new_values['totime'] = request_totime
                        new_values['absence_flg'] = False
                    obj = SubmitShift(**new_values)
                    obj.save()
        return redirect('submit_shifts')
    
def export(request, year, month):
    days = common.getCalendarDays(year, month)
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(year) + '年' + str(month) + '月'
    # 指定の年月の提出されたシフトデータを全て取得
    submited_shifts = {}
    staffs = Staff.objects.all()
    for staff in staffs:
        target_shifts = SubmitShift.objects.filter(year=year, month=month, staff_id=staff).order_by('day')
        submited_shifts[staff.staff_name] = target_shifts

    # 書き込み処理
    # 日付
    x = 2
    y = 1
    for day, values in days.items():
        ws.cell(row=y, column=x).value = str(month) + '/' + str(day) + '(' + values['day_of_week'] + ')'
        if values['holiday_flg'] or day == 28:
            ws.cell(row=y, column=x).font = openpyxl.styles.fonts.Font(color='FF0000')
        x = x + 1
    # シフト
    y = 2
    for name, shifts in submited_shifts.items():
        x = 1
        ws.cell(row=y, column=x).value = name
        ws.cell(row=y, column=x).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")
        x = x + 1
        for shift in shifts:
            ws.cell(row=y, column=x).value = 'x' if shift.absence_flg else str(shift.fromtime) + '~' + str(shift.totime)
            ws.cell(row=y, column=x).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")
            x = x + 1
        y = y + 1  
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=submitted_shift.xlsx'
    return response