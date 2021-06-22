from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from .models import  CompleteShift 
from django.contrib.auth.models import User
from staff.models import Staff
from common import auto_create
from common import common
import io
import openpyxl
from django.http import HttpResponse

# Create your views here.

class CreateShiftView(TemplateView):
    template_name = 'shifts/shift_complete.html'
    def get(self, request):
        shift = CompleteShift.objects.filter(year=auto_create.year, month=auto_create.month)
        status = Staff.objects.get(staff_id=request.user)
        context = {'year':auto_create.year, 'month':auto_create.month, 'shift':shift, 'days':auto_create.days, 'status':status}
        return render(request, self.template_name, context)
    
    def post(self, request):
        auto_create.auto_create()
        return redirect('create_shift')

def export(request, year, month):
    days = common.getCalendarDays(year, month)
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(year) + '年' + str(month) + '月'

    # 書き込み処理
    x = 2
    ws.cell(row=2, column=1).value = 'Flyer'
    ws.cell(row=11, column=1).value = 'Counter'
    ws.cell(row=28, column=1).value = 'Kitchen'
    for day, values in days.items():
        y = 1
        ws.cell(row=y, column=x).value = str(month) + '/' + str(day) + '(' + values['day_of_week'] + ')'
        if values['holiday_flg'] or day == 28:
            ws.cell(row=y, column=x).font = openpyxl.styles.fonts.Font(color='FF0000')
        y = y + 1
        shifts = CompleteShift.objects.filter(year=year, month=month, day=day)[0]
        if shifts.flyer_shift.am_1 is not None:
            fly_am1 = shifts.flyer_shift.am_1
            write_excel(ws, y, x, fly_am1)
        y = y + 2 
        if shifts.flyer_shift.am_2 is not None:
            fly_am2 = shifts.flyer_shift.am_2
            write_excel(ws, y, x, fly_am2)
        y = y + 2  
        if shifts.flyer_shift.pm_1 is not None:
            fly_pm1 = shifts.flyer_shift.pm_1
            write_excel(ws, y, x, fly_pm1)
        y = y + 2
        if shifts.flyer_shift.pm_2 is not None:
            fly_pm2 = shifts.flyer_shift.pm_2
            write_excel(ws, y, x, fly_pm2)
        y = y + 3
        if shifts.counter_shift.am_1 is not None:
            counter_am1 = shifts.counter_shift.am_1
            write_excel(ws, y, x, counter_am1)
        y = y + 2
        if shifts.counter_shift.am_2 is not None:
            counter_am2 = shifts.counter_shift.am_2
            write_excel(ws, y, x, counter_am2)
        y = y + 2 
        if shifts.counter_shift.am_3 is not None:
            counter_am3 = shifts.counter_shift.am_3
            write_excel(ws, y, x, counter_am3)
        y = y + 2    
        if shifts.counter_shift.am_4 is not None:
            counter_am4 = shifts.counter_shift.am_4
            write_excel(ws, y, x, counter_am4)
        y = y + 2
        if shifts.counter_shift.am_5 is not None:
            counter_am5 = shifts.counter_shift.am_5
            write_excel(ws, y, x, counter_am5)
        y = y + 2 
        if shifts.counter_shift.pm_1 is not None:
            counter_pm1 = shifts.counter_shift.pm_1
            write_excel(ws, y, x, counter_pm1)
        y = y + 2 
        if shifts.counter_shift.pm_2 is not None:
            counter_pm2 = shifts.counter_shift.pm_2
            write_excel(ws, y, x, counter_pm2)
        y = y + 2
        if shifts.counter_shift.pm_3 is not None:
            counter_pm3 = shifts.counter_shift.pm_3
            write_excel(ws, y, x, counter_pm3)
        y = y + 3
        if shifts.kitchen_shift.am_1 is not None:
            kit_am1 = shifts.kitchen_shift.am_1
            write_excel(ws, y, x, kit_am1)
        y = y + 2 
        if shifts.kitchen_shift.am_2 is not None:
            kit_am2 = shifts.kitchen_shift.am_2
            write_excel(ws, y, x, kit_am2)
        y = y + 2  
        if shifts.kitchen_shift.pm_1 is not None:
            kit_pm1 = shifts.kitchen_shift.pm_1
            write_excel(ws, y, x, kit_pm1)
        y = y + 2 
        if shifts.kitchen_shift.pm_2 is not None:
            kit_pm2 = shifts.kitchen_shift.pm_2
            write_excel(ws, y, x, kit_pm2)
        x = x + 1 
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=submitted_shift.xlsx'
    return response


def write_excel(sheet, y, x, position):
    sheet.cell(row=y, column=x).value = str(position.modify_fromtime) + ':00~' + str(position.modify_totime) + ':00'
    if position.submit_id is not None:
        sheet.cell(row=y+1, column=x).value = position.submit_id.staff_id.staff_name
        sheet.cell(row=y+1, column=x).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")

    else:
        sheet.cell(row=y+1, column=x).value = '(　　)'
        fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='d3d3d3')
        sheet.cell(row=y+1, column=x).fill = fill
        sheet.cell(row=y+1, column=x).alignment = openpyxl.styles.Alignment(horizontal="centerContinuous")